import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
CORE_PATH = ROOT / "03_APPS_SCRIPT_KOD" / "tesbih_kuyusu_v6_5_ultra_operasyon_core.gs"
DEFAULT_XLSX = Path(r"C:\Users\emrah\Downloads\TESBIH_KUYUSU_MASTER_SHEET (20).xlsx")


def norm(value):
    text = str(value or "").strip().translate({0x0131: "i", 0x0130: "i"}).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def cell_text(value):
    return "" if value is None else str(value).strip()


def load_code_headers():
    node_exe = os.environ.get("NODE_EXE", "node")
    script = """
const fs = require("fs");
const vm = require("vm");
const corePath = process.argv[1];
const code = fs.readFileSync(corePath, "utf8");
const sandbox = { console, Date, JSON, Math, Number, String, Object, Array, RegExp, encodeURIComponent };
vm.createContext(sandbox);
vm.runInContext(code, sandbox, { filename: "tesbih_kuyusu_v6_5_ultra_operasyon_core.gs" });
process.stdout.write(JSON.stringify(sandbox.TK6.HEADERS));
"""
    result = subprocess.run(
        [node_exe, "-e", script, str(CORE_PATH)],
        check=True,
        text=True,
        encoding="utf-8",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    return json.loads(result.stdout)


def sheet_headers(workbook, sheet_name):
    values = [cell_text(cell.value) for cell in workbook[sheet_name][1]]
    while values and not values[-1]:
        values.pop()
    return values


def records(workbook, sheet_name):
    header_cells = [cell_text(cell.value) for cell in workbook[sheet_name][1]]
    index = {norm(name): i for i, name in enumerate(header_cells) if name}
    output = []
    for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
        values = [cell_text(value) for value in row]
        if not any(values):
            continue
        output.append({name: values[i] if i < len(values) else "" for name, i in index.items()})
    return output


def fail(message, details=None):
    payload = {"ok": False, "error": message}
    if details is not None:
        payload["details"] = details
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    sys.exit(1)


def main():
    xlsx_path = Path(os.environ.get("TESBIH_REFERENCE_XLSX", str(DEFAULT_XLSX)))
    if not xlsx_path.exists():
        fail("Referans XLSX bulunamadı", str(xlsx_path))

    workbook = load_workbook(xlsx_path, data_only=False, read_only=False)
    code_headers = load_code_headers()
    required = {
        "settings": "01_AYARLAR",
        "open": "03_ACIK_SIPARISLER",
        "items": "04_URUN_KALEMLERI",
        "payments": "05_ODEMELER",
        "invoiceGroups": "06_FATURA_GRUPLARI",
        "parasut": "07_PARASUT_FATURA",
        "cargo": "08_KARGO_PAKETLERI",
        "finance808": "10_808_FINANS_ONIZLEME",
        "ebelge": "11_EBELGE_ISTISNA",
        "control": "12_KONTROL_MERKEZI",
        "dictionary": "13_VERI_SOZLUGU",
        "bank": "14_BANKA_HAREKETLERI",
        "addressMemory": "15_MUSTERI_ADRESLERI",
    }

    missing_sheets = [sheet for sheet in required.values() if sheet not in workbook.sheetnames]
    if missing_sheets:
        fail("Zorunlu sayfa eksik", missing_sheets)

    header_mismatches = {}
    for key, sheet_name in required.items():
        actual = sheet_headers(workbook, sheet_name)
        expected = code_headers.get(key, [])
        if actual != expected:
            actual_norm = {norm(item) for item in actual}
            expected_norm = {norm(item) for item in expected}
            header_mismatches[sheet_name] = {
                "sheet_extra": [item for item in actual if norm(item) not in expected_norm],
                "code_missing_in_sheet": [item for item in expected if norm(item) not in actual_norm],
                "order_mismatch": actual_norm == expected_norm and actual != expected,
                "actual": actual,
                "expected": expected,
            }
    if header_mismatches:
        fail("Sheet referans kolon sözleşmesi kod HEADERS ile eşleşmiyor", header_mismatches)

    open_ids = {row.get("acik_siparis_id", "") for row in records(workbook, "03_ACIK_SIPARISLER")}
    open_ids.discard("")
    if not open_ids:
        fail("03_ACIK_SIPARISLER içinde Açık_Sipariş_ID bulunamadı")

    link_errors = {}
    for sheet_name in [
        "04_URUN_KALEMLERI",
        "05_ODEMELER",
        "06_FATURA_GRUPLARI",
        "07_PARASUT_FATURA",
        "08_KARGO_PAKETLERI",
        "10_808_FINANS_ONIZLEME",
        "11_EBELGE_ISTISNA",
    ]:
        ids = [row.get("acik_siparis_id", "") for row in records(workbook, sheet_name) if row.get("acik_siparis_id", "")]
        missing = sorted({open_id for open_id in ids if open_id not in open_ids})
        if missing:
            link_errors[sheet_name] = missing
    if link_errors:
        fail("Bağlı sayfalarda 03 içinde bulunmayan Açık_Sipariş_ID var", link_errors)

    dictionary_rows = []
    for row in workbook["13_VERI_SOZLUGU"].iter_rows(min_row=2, values_only=True):
        values = [cell_text(value) for value in row]
        if any(values):
            dictionary_rows.append(values)
    dictionary_pairs = {(row[0], row[1]) for row in dictionary_rows if len(row) > 1 and row[0] and row[1]}
    dictionary_missing = []
    for sheet_name in required.values():
        if sheet_name == "13_VERI_SOZLUGU":
            continue
        for header in sheet_headers(workbook, sheet_name):
            if (sheet_name, header) not in dictionary_pairs:
                dictionary_missing.append({"Sayfa": sheet_name, "Kolon": header})
    if dictionary_missing:
        fail("13_VERI_SOZLUGU gerçek kolonları kapsamıyor", dictionary_missing)

    finance_errors = []
    for index, row in enumerate(records(workbook, "10_808_FINANS_ONIZLEME"), start=2):
        raw = row.get("fark", "")
        if not raw:
            continue
        try:
            value = float(str(raw).replace(".", "").replace(",", "."))
        except ValueError:
            finance_errors.append({"row": index, "Fark": raw, "reason": "parse"})
            continue
        if abs(value) > 1e-9:
            finance_errors.append({"row": index, "Fark": raw, "reason": "nonzero"})
    if finance_errors:
        fail("10_808_FINANS_ONIZLEME Fark alanı sıfır olmayan kayıt içeriyor", finance_errors)

    payment_status = Counter(row.get("teyit_durumu", "") for row in records(workbook, "05_ODEMELER"))
    bank_rows = len(records(workbook, "14_BANKA_HAREKETLERI"))
    required_columns = {
        "07_PARASUT_FATURA": ["Payload_JSON", "Response_JSON", "Gönderim_Tarihi"],
        "08_KARGO_PAKETLERI": [
            "Navlungo_Post_Number",
            "Navlungo_Tracking_URL",
            "Navlungo_Barcode_URL",
            "Navlungo_Status",
            "Barkod_Yazdirildi_Mi",
            "Barkod_Yazdirma_Sonucu",
            "Barkod_Yazdirma_Hata",
        ],
    }
    missing_required_columns = {}
    for sheet_name, names in required_columns.items():
        actual_norm = {norm(header) for header in sheet_headers(workbook, sheet_name)}
        missing = [name for name in names if norm(name) not in actual_norm]
        if missing:
            missing_required_columns[sheet_name] = missing
    if missing_required_columns:
        fail("Zorunlu referans kolonları eksik", missing_required_columns)

    result = {
        "ok": True,
        "xlsx": str(xlsx_path),
        "sha256": hashlib.sha256(xlsx_path.read_bytes()).hexdigest().upper(),
        "sheet_count": len(workbook.sheetnames),
        "open_id_count": len(open_ids),
        "data_dictionary_rows": len(dictionary_rows),
        "payment_status": dict(payment_status),
        "bank14_rows": bank_rows,
        "checked": [
            "Sheet/code header parity",
            "ID link integrity",
            "13_VERI_SOZLUGU coverage",
            "10_808 Fark zero",
            "07/08 required columns",
            "05 payment status read-only verification",
            "14 bank movement empty verification",
        ],
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
